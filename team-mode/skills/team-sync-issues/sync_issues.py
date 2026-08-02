#!/usr/bin/env python3
"""
Sync GitHub issues with an xlsx tracker.

GH is authoritative for: Issue #, Title, Labels, URL, Status, created_date, modified_date.
Human is authoritative for: Priority, Type, Notes (preserved across syncs).

Usage:
    python sync_issues.py
    python sync_issues.py --xlsx PATH
    python sync_issues.py --dry-run
    python sync_issues.py --repo OWNER/REPO
"""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    "Issue #", "Priority", "Type", "Status",
    "Title", "Labels", "URL", "Notes",
    "created_date", "modified_date",
]

DEFAULT_COL_WIDTHS = {
    "Issue #": 8.5, "Priority": 10.0, "Type": 14.7, "Status": 24.2,
    "Title": 91.8, "Labels": 54.0, "URL": 47.3, "Notes": 62.0,
    "created_date": 13.8, "modified_date": 14.8,
}

# Status color coding (only applied when CREATING a new file; preserved when updating existing)
STATUS_FILL = {
    "READY":         PatternFill("solid", start_color="E7F5E8"),  # light green
    "IN PROGRESS":   PatternFill("solid", start_color="FFF4D6"),  # light amber
    "IN REVIEW":     PatternFill("solid", start_color="DEEBFF"),  # light blue
    "TESTING":       PatternFill("solid", start_color="DEEBFF"),
    "PR FOLLOW-UP":  PatternFill("solid", start_color="FFE8D6"),  # peach
    "BLOCKED":       PatternFill("solid", start_color="FAD1CF"),  # light red
    "CLOSED":        PatternFill("solid", start_color="EEEEEE"),  # grey
    "MISSING":       PatternFill("solid", start_color="FAD1CF"),
}


# ---------- GitHub ----------

def gh_json(args: list[str]) -> list | dict | None:
    """Run a gh command expecting JSON output."""
    try:
        result = subprocess.run(
            ["gh"] + args,
            capture_output=True, text=True, check=True,
        )
    except FileNotFoundError:
        sys.exit("error: `gh` CLI not found. Install https://cli.github.com and run `gh auth login`.")
    except subprocess.CalledProcessError as e:
        sys.exit(f"error: gh command failed: {' '.join(args)}\n{e.stderr}")
    return json.loads(result.stdout) if result.stdout.strip() else None


def fetch_gh_issues(repo: str | None) -> dict[int, dict]:
    """Fetch open + recently-closed issues. Returns {number: issue_dict}."""
    base = ["--repo", repo] if repo else []
    fields = "number,title,labels,url,createdAt,updatedAt,state,closedAt"

    open_issues = gh_json(["issue", "list", "--state", "open", "--limit", "500",
                           "--json", fields] + base) or []
    closed_issues = gh_json(["issue", "list", "--state", "closed", "--limit", "100",
                             "--json", fields] + base) or []

    by_num = {}
    for issue in open_issues:
        issue["state"] = issue["state"].lower()
        by_num[issue["number"]] = issue
    for issue in closed_issues:
        issue["state"] = issue["state"].lower()
        by_num[issue["number"]] = issue
    return by_num


# ---------- Inference ----------

TITLE_PREFIX_TO_TYPE = [
    ("fix:", "BUG"),
    ("feat:", "FEATURE"),
    ("feature:", "FEATURE"),
    ("chore:", "CHORE"),
    ("refactor:", "CHORE"),
    ("test:", "CHORE"),
    ("perf:", "CHORE"),
    ("ci:", "CHORE"),
    ("build:", "CHORE"),
    ("docs:", "DOCS"),
    ("doc:", "DOCS"),
    ("data:", "DATA"),
]

LABEL_TO_TYPE = [
    ("BUG", "BUG"),
    ("DOCUMENTATION", "DOCS"),
    ("DOCS", "DOCS"),
    ("ENHANCEMENT", "FEATURE"),
    ("FEATURE", "FEATURE"),
    ("CHORE", "CHORE"),
    ("data-quality", "DATA"),
    ("data-mapping", "DATA"),
]


def infer_type(title: str, labels: list[str]) -> str:
    t = (title or "").lower().strip()
    for prefix, type_ in TITLE_PREFIX_TO_TYPE:
        if t.startswith(prefix):
            return type_
    label_set = {l.lower() for l in labels}
    for label, type_ in LABEL_TO_TYPE:
        if label.lower() in label_set:
            return type_
    return "CHORE"


def infer_priority(labels: list[str]) -> str:
    label_set = {l.lower() for l in labels}
    if label_set & {"urgent", "high-priority", "p0", "p1"}:
        return "1 - HIGH"
    if label_set & {"low-priority", "p3"}:
        return "3 - LOW"
    return "2 - MEDIUM"


def infer_status(labels: list[str], state: str) -> str:
    if state == "closed":
        return "CLOSED"
    label_set = {l.lower() for l in labels}
    for label, status in [
        ("in-progress", "IN PROGRESS"),
        ("in-review", "IN REVIEW"),
        ("testing", "TESTING"),
        ("pr-follow-up", "PR FOLLOW-UP"),
        ("blocked", "BLOCKED"),
    ]:
        if label in label_set:
            return status
    return "READY"


def format_date(iso_str: str | None) -> str:
    return iso_str[:10] if iso_str else ""


# ---------- xlsx I/O ----------

def load_existing(path: Path) -> tuple[dict[int, dict], list[str], dict[int, int]]:
    """Return (rows_by_num, headers, row_idx_by_num). Empty dict if file missing."""
    if not path.exists():
        return {}, list(COLUMNS), {}

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = {}
    row_idx_by_num = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = dict(zip(headers, row))
        raw = row_dict.get("Issue #")
        if raw is None:
            continue
        try:
            num = int(str(raw).replace("#", "").strip())
        except (ValueError, TypeError):
            continue
        rows[num] = row_dict
        row_idx_by_num[num] = row_idx
    return rows, headers, row_idx_by_num


def create_new_workbook(path: Path) -> None:
    """Create a fresh workbook with defaults (only when file is missing)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Open Issues"
    ws.append(COLUMNS)
    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = DEFAULT_COL_WIDTHS.get(header, 14)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(path)


def build_gh_row(num: int, gh_issue: dict) -> dict:
    labels = [lbl["name"] for lbl in gh_issue.get("labels", [])]
    return {
        "Issue #": str(num),
        "Priority": infer_priority(labels),
        "Type": infer_type(gh_issue["title"], labels),
        "Status": infer_status(labels, gh_issue["state"]),
        "Title": gh_issue["title"],
        "Labels": ", ".join(labels) if labels else None,
        "URL": gh_issue["url"],
        "Notes": None,
        "created_date": format_date(gh_issue.get("createdAt")),
        "modified_date": format_date(gh_issue.get("updatedAt")),
    }


def volatile_fields_changed(existing: dict, gh_issue: dict) -> bool:
    labels = [lbl["name"] for lbl in gh_issue.get("labels", [])]
    new_status = infer_status(labels, gh_issue["state"])
    new_labels_str = ", ".join(labels) if labels else None
    return (
        (existing.get("Title") or "") != (gh_issue["title"] or "")
        or (existing.get("Labels") or "") != (new_labels_str or "")
        or (existing.get("Status") or "") != new_status
        or (existing.get("URL") or "") != gh_issue["url"]
        or (existing.get("modified_date") or "") != format_date(gh_issue.get("updatedAt"))
    )


def compute_diff(gh_issues: dict[int, dict], existing: dict[int, dict]) -> dict:
    """Classify each issue."""
    add, update, close, unchanged, missing = [], [], [], [], []

    for num, gh_issue in gh_issues.items():
        state = gh_issue["state"]
        in_xlsx = num in existing
        if state == "open" and not in_xlsx:
            add.append(num)
        elif state == "open" and in_xlsx:
            if volatile_fields_changed(existing[num], gh_issue):
                update.append(num)
            else:
                unchanged.append(num)
        elif state == "closed" and in_xlsx:
            close.append(num)
        # closed + not in xlsx: ignore (was never tracked)

    # Anything in xlsx that GH has no record of at all
    for num in existing:
        if num not in gh_issues:
            missing.append(num)

    return {"add": add, "update": update, "close": close,
            "unchanged": unchanged, "missing": missing}


def apply_updates(path: Path, gh_issues: dict[int, dict], existing_rows: dict[int, dict],
                  headers: list[str], row_idx_by_num: dict[int, int], diff: dict) -> None:
    """Apply updates to the workbook in place. Preserves cell formatting."""
    wb = load_workbook(path)
    ws = wb.active

    # Map header name to column index (1-based)
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    # --- UPDATE: modify volatile fields in place ---
    for num in diff["update"]:
        row_i = row_idx_by_num[num]
        gh_issue = gh_issues[num]
        labels = [lbl["name"] for lbl in gh_issue.get("labels", [])]
        updates = {
            "Title": gh_issue["title"],
            "Labels": ", ".join(labels) if labels else None,
            "URL": gh_issue["url"],
            "Status": infer_status(labels, gh_issue["state"]),
            "modified_date": format_date(gh_issue.get("updatedAt")),
        }
        for col_name, value in updates.items():
            if col_name in col_idx:
                ws.cell(row=row_i, column=col_idx[col_name], value=value)

    # --- ADD: append new rows at the bottom ---
    # Copy style from the last data row if available
    template_row_i = None
    if ws.max_row >= 2:
        template_row_i = ws.max_row

    for num in diff["add"]:
        row = build_gh_row(num, gh_issues[num])
        new_row_i = ws.max_row + 1
        for col_name, value in row.items():
            if col_name in col_idx:
                cell = ws.cell(row=new_row_i, column=col_idx[col_name], value=value)
                # Copy font/alignment from template row (not the fill — keep it default)
                if template_row_i and template_row_i != new_row_i:
                    src = ws.cell(row=template_row_i, column=col_idx[col_name])
                    if src.font:
                        cell.font = copy(src.font)
                    if src.alignment:
                        cell.alignment = copy(src.alignment)

    # --- CLOSE + MISSING: move to "Recently Closed" sheet ---
    if diff["close"] or diff["missing"]:
        closed_ws = wb["Recently Closed"] if "Recently Closed" in wb.sheetnames else wb.create_sheet("Recently Closed")
        if closed_ws.max_row == 1 and all(c.value is None for c in closed_ws[1]):
            # Fresh sheet — add headers
            for i, h in enumerate(headers, start=1):
                hc = closed_ws.cell(row=1, column=i, value=h)
                src = ws.cell(row=1, column=i)
                if src.font:
                    hc.font = copy(src.font)
            closed_ws.freeze_panes = "A2"
            for i, h in enumerate(headers, start=1):
                closed_ws.column_dimensions[get_column_letter(i)].width = \
                    ws.column_dimensions[get_column_letter(i)].width or DEFAULT_COL_WIDTHS.get(h, 14)

        # Collect rows to move (reverse-sort indices so deletion doesn't shift).
        to_move = sorted(
            [(row_idx_by_num[n], n, "CLOSED") for n in diff["close"] if n in row_idx_by_num] +
            [(row_idx_by_num[n], n, "MISSING") for n in diff["missing"] if n in row_idx_by_num],
            reverse=True,
        )
        for row_i, num, new_status in to_move:
            vals = []
            for i in range(1, len(headers) + 1):
                vals.append(ws.cell(row=row_i, column=i).value)
            # Overwrite Status column with CLOSED/MISSING
            if "Status" in col_idx:
                vals[col_idx["Status"] - 1] = new_status
            # Update modified_date from GH if we have it
            if num in gh_issues and "modified_date" in col_idx:
                closed_at = gh_issues[num].get("closedAt") or gh_issues[num].get("updatedAt")
                vals[col_idx["modified_date"] - 1] = format_date(closed_at)
            closed_ws.append(vals)
        # Delete from main sheet
        for row_i, _, _ in to_move:
            ws.delete_rows(row_i, 1)

    wb.save(path)


# ---------- Reporting ----------

def print_diff(diff: dict, gh_issues: dict, existing: dict, dry: bool) -> None:
    prefix = "[DRY-RUN] " if dry else ""
    print(f"\n{prefix}Sync summary")
    print("=" * 60)
    print(f"  ADD (new in GH):         {len(diff['add']):4d}")
    print(f"  UPDATE (fields changed): {len(diff['update']):4d}")
    print(f"  CLOSE (closed in GH):    {len(diff['close']):4d}")
    print(f"  MISSING (removed in GH): {len(diff['missing']):4d}")
    print(f"  UNCHANGED:               {len(diff['unchanged']):4d}")
    print()
    for tag, nums in [("ADD", diff["add"]), ("UPDATE", diff["update"]),
                      ("CLOSE", diff["close"]), ("MISSING", diff["missing"])]:
        if not nums:
            continue
        print(f"--- {tag} ---")
        for num in sorted(nums):
            title = ""
            if num in gh_issues:
                title = gh_issues[num].get("title", "")
            elif num in existing:
                title = existing[num].get("Title", "") or ""
            print(f"  #{num}: {title[:80]}")
        print()


# ---------- Backup ----------

def make_backup(path: Path) -> Path | None:
    if not path.exists():
        return None
    backup_dir = path.parent / "backups"
    backup_dir.mkdir(exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    dest = backup_dir / f"{path.stem}_{ts}{path.suffix}"
    shutil.copy2(path, dest)
    return dest


# ---------- Main ----------

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", default="tmp/_OPEN_ISSUES.xlsx",
                        help="Path to tracker xlsx (default: tmp/_OPEN_ISSUES.xlsx)")
    parser.add_argument("--repo", default=None,
                        help="GitHub repo OWNER/REPO (default: derived from git remote)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show diff without writing")
    args = parser.parse_args()

    path = Path(args.xlsx).resolve()
    print(f"xlsx:   {path}")
    print(f"repo:   {args.repo or '(from git remote)'}")
    print(f"mode:   {'DRY-RUN' if args.dry_run else 'APPLY'}")

    # 1. Backup if file exists
    if not args.dry_run:
        backup = make_backup(path)
        if backup:
            print(f"backup: {backup}")

    # 2. Create file if missing
    created = False
    if not path.exists():
        print(f"note:   file missing, creating fresh: {path}")
        if not args.dry_run:
            path.parent.mkdir(parents=True, exist_ok=True)
            create_new_workbook(path)
        created = True

    # 3. Load existing
    existing_rows, headers, row_idx_by_num = load_existing(path) if path.exists() else ({}, list(COLUMNS), {})

    # 4. Fetch GH
    print("fetching GH issues…")
    gh_issues = fetch_gh_issues(args.repo)
    print(f"  fetched {len(gh_issues)} issues (open + recently closed)")

    # 5. Diff
    diff = compute_diff(gh_issues, existing_rows)
    print_diff(diff, gh_issues, existing_rows, args.dry_run)

    total_changes = len(diff["add"]) + len(diff["update"]) + len(diff["close"]) + len(diff["missing"])
    if total_changes == 0 and not created:
        print("nothing to do.")
        return 0

    # 6. Apply
    if args.dry_run:
        print("DRY-RUN: no changes written.")
        return 0

    if not path.exists():
        # was --dry-run'd file creation path; shouldn't hit here
        create_new_workbook(path)

    apply_updates(path, gh_issues, existing_rows, headers, row_idx_by_num, diff)
    print(f"\nwrote: {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
