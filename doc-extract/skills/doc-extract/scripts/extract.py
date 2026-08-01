#!/usr/bin/env python3
"""Universal structural document extraction — one entry point, any format.

Usage:
  extract.py file.{pdf,docx,xlsx,pptx,csv,tsv,txt,md,html,epub,rtf,odt}
  extract.py --json file.pdf     # tagged elements (pdf only; others emit markdown)
  extract.py --selftest

Routes each format to the tool that preserves its structure:
  pdf              -> pdf_structure.py (headings, tables, forms, reading order)
  docx/html/epub/  -> pandoc -t gfm (these formats already store structure)
    rtf/odt
  xlsx             -> openpyxl, one GFM table per sheet
  pptx             -> stdlib zip/XML, text per slide (no dependency needed)
  csv/tsv          -> stdlib csv -> GFM table
  txt/md/json      -> passthrough (already machine-readable)
Output is Markdown on stdout, ready for chunking or direct reading.
"""
import csv
import re
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

MAX_ROWS = 500  # ponytail: hard cap per sheet/csv; truncation is announced, not silent

PANDOC_FORMATS = {".docx", ".html", ".htm", ".epub", ".rtf", ".odt"}
PASSTHROUGH = {".txt", ".md", ".markdown", ".json", ".log", ".rst"}


def rows_to_gfm(rows):
    def clean(c):
        return "" if c is None else str(c).replace("\n", " ").replace("|", "\\|").strip()
    rows = [r for r in rows if any(c not in (None, "") for c in r)]
    if not rows:
        return "_empty_"
    width = max(len(r) for r in rows)
    rows = [list(r) + [""] * (width - len(r)) for r in rows]
    out = ["| " + " | ".join(clean(c) for c in rows[0]) + " |",
           "|" + "---|" * width]
    body = rows[1:MAX_ROWS + 1]
    out += ["| " + " | ".join(clean(c) for c in r) + " |" for r in body]
    if len(rows) - 1 > MAX_ROWS:
        out.append(f"\n_[truncated: {len(rows) - 1 - MAX_ROWS} more rows not shown]_")
    return "\n".join(out)


def extract_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    parts = []
    for ws in wb.worksheets:
        parts.append(f"## Sheet: {ws.title}\n")
        parts.append(rows_to_gfm([[c for c in row] for row in ws.iter_rows(values_only=True)]))
        parts.append("")
    return "\n".join(parts)


def extract_pptx(path):
    # a .pptx is a zip; slide text lives in <a:t> elements — no library needed
    parts = []
    with zipfile.ZipFile(path) as z:
        slides = sorted(
            (n for n in z.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", n)),
            key=lambda n: int(re.search(r"\d+", n).group()))
        for i, name in enumerate(slides, 1):
            xml = z.read(name).decode("utf-8", errors="replace")
            texts = re.findall(r"<a:t>(.*?)</a:t>", xml, re.S)
            parts.append(f"## Slide {i}\n")
            parts.append("\n".join(t.strip() for t in texts if t.strip()) or "_no text_")
            parts.append("")
    return "\n".join(parts)


def extract_csv(path, delimiter):
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        return rows_to_gfm(list(csv.reader(f, delimiter=delimiter)))


def extract_docx_stdlib(path):
    # a .docx is a zip; paragraph text lives in <w:t> runs inside <w:p> blocks.
    # Loses tables/heading levels — fallback only, used when pandoc is absent.
    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml").decode("utf-8", errors="replace")
    paras = []
    for p in re.findall(r"<w:p[ >].*?</w:p>", xml, re.S):
        text = "".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", p, re.S))
        if text.strip():
            paras.append(text.strip())
    return "\n\n".join(paras)


def extract_pandoc(path):
    if not shutil.which("pandoc"):
        if path.suffix.lower() == ".docx":
            return ("<!-- pandoc not installed: stdlib fallback, "
                    "tables/headings flattened -->\n\n" + extract_docx_stdlib(path))
        sys.exit(f"pandoc is required for {path.suffix} but is not installed "
                 f"(e.g. `brew install pandoc` or `apt-get install pandoc`).")
    r = subprocess.run(["pandoc", "-t", "gfm", str(path)],
                       capture_output=True, text=True)
    if r.returncode != 0:
        sys.exit(f"pandoc failed on {path}: {r.stderr.strip()}")
    return r.stdout


def extract(path: Path, as_json=False) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        sys.path.insert(0, str(Path(__file__).parent))
        try:
            import fitz
        except ImportError:
            sys.exit("PyMuPDF is required for PDFs but not installed: "
                     "run `pip install pymupdf` and re-run.")
        import pdf_structure
        import json as _json
        doc = fitz.open(path)
        els = pdf_structure.extract(doc)
        return _json.dumps(els, indent=2) if as_json else pdf_structure.to_markdown(els)
    if ext in PANDOC_FORMATS:
        return extract_pandoc(path)
    if ext == ".xlsx":
        return extract_xlsx(path)
    if ext == ".pptx":
        return extract_pptx(path)
    if ext in (".csv", ".tsv"):
        return extract_csv(path, "\t" if ext == ".tsv" else ",")
    if ext in PASSTHROUGH:
        return path.read_text(encoding="utf-8", errors="replace")
    if ext == ".doc" or ext == ".xls" or ext == ".ppt":
        sys.exit(f"{ext} is a legacy binary format; convert it first "
                 f"(e.g. `soffice --convert-to {ext}x`) or ask how to proceed.")
    sys.exit(f"unsupported extension: {ext}")


def selftest():
    import tempfile
    from openpyxl import Workbook
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        # xlsx
        wb = Workbook()
        ws = wb.active
        ws.title = "Q1"
        ws.append(["Region", "Sales"])
        ws.append(["West", 500])
        wb.save(td / "t.xlsx")
        md = extract(td / "t.xlsx")
        assert "## Sheet: Q1" in md and "| West | 500 |" in md, md
        # csv
        (td / "t.csv").write_text("a,b\n1,2\n")
        md = extract(td / "t.csv")
        assert "| a | b |" in md and "| 1 | 2 |" in md, md
        # pptx (minimal hand-built zip)
        slide = ('<p:sld xmlns:a="urn:x" xmlns:p="urn:y"><a:t>Hello deck</a:t></p:sld>')
        with zipfile.ZipFile(td / "t.pptx", "w") as z:
            z.writestr("ppt/slides/slide1.xml", slide)
        md = extract(td / "t.pptx")
        assert "## Slide 1" in md and "Hello deck" in md, md
        # passthrough
        (td / "t.md").write_text("# hi\n")
        assert extract(td / "t.md") == "# hi\n"
        # docx stdlib fallback (minimal hand-built zip)
        docx_xml = ('<w:document xmlns:w="urn:w"><w:body>'
                    '<w:p><w:r><w:t>Hello doc</w:t></w:r></w:p>'
                    '</w:body></w:document>')
        with zipfile.ZipFile(td / "t.docx", "w") as z:
            z.writestr("word/document.xml", docx_xml)
        assert "Hello doc" in extract_docx_stdlib(td / "t.docx")
    print("selftest OK")


def main():
    args = sys.argv[1:]
    if "--selftest" in args:
        return selftest()
    as_json = "--json" in args
    args = [a for a in args if a != "--json"]
    if not args:
        sys.exit(__doc__)
    print(extract(Path(args[0]), as_json=as_json))


if __name__ == "__main__":
    main()
